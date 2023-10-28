"""

restaurant_list 에는 공공 api를 이용해 추출한 부산 광역시 금정구 인근의 식당 목록이 저장되어 있다.

해당 py 파일은
restaurant_list 에 있는 식당 목록 중 YOLOv8 학습 데이터가 있는
햄버거, 파스타, 피자, 치킨, 스시 5개 카텍고리에 해당되는 식당 목록을 추출하여
해당 식당의 이름과 카테고리를 csv 형식으로 저장한다.

"""


import openpyxl
import csv

# Excel 파일 열기
wb = openpyxl.load_workbook('restaurant_list.xlsx')

# 워크시트 선택
sheet = wb.active  # 기본 워크시트 선택
# 또는 워크시트 이름으로 선택: sheet = wb['Sheet1']

name_index = None
cate_index = None

for cell in sheet[1]:
    if cell.value == "이름":
        name_index = cell.column
    elif cell.value == "항목":
        cate_index = cell.column
        
# "이름" 열의 데이터 읽기     
res_names = []
if name_index is not None:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name_value = row[name_index - 1]
        res_names.append(name_value if name_value is not None else '')

# "카테고리" 열의 데이터 읽기
res_cates = []
if cate_index is not None:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cate_value = row[cate_index - 1]
        res_cates.append(cate_value if cate_value is not None else '')

# 워크북 닫기
wb.close()

category_to_food = {
    "닭요리": "chicken",
    "치킨": "chicken",
    "피자": "pizza",
    "이탈리안": "pizza",
    "패스트푸드": "pizza",
    "참치": "sushi",
    "일식": "sushi",
    "해물": "sushi",
    "초밥": "sushi",
    "회": "sushi",
    "패스트푸드": "burger",
    "햄버거": "burger",
    "이탈리안": "pasta",
    "양식": "pasta",
}

with open('target_list.csv', 'w', newline='', encoding='utf-8') as csvfile:
    csvwriter = csv.writer(csvfile)
    csvwriter.writerow(["이름", "카테고리"])  # 열 이름을 추가

    # name과 cate 리스트 데이터를 행으로 저장
    for i in range(len(res_names)):
        if res_names[i] == '':
            continue
        food_categories = set()  # 중복을 허용하는 집합 생성
        for category, food in category_to_food.items():
            if category in res_cates[i]:
                food_categories.add(food)  # 중복되지 않는 음식 종류 추가
        for food in food_categories:
            csvwriter.writerow([res_names[i], food])