import openpyxl
from openpyxl import Workbook
from os.path import dirname, realpath, exists

# Open Brand
org_path = dirname(realpath(__file__).replace("\\", "/"))

def get_category_data(excel_file_path):
    excel_full_path = '/'.join([org_path, excel_file_path])
    wb = openpyxl.load_workbook(excel_full_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data



restaurant_data = get_category_data('data.xlsx')

excel_full_path = '/'.join([org_path, 'search_kakao.xlsx'])
if exists(excel_full_path):
    wb = openpyxl.load_workbook(excel_full_path)
else:
    wb = Workbook()
sheet = wb.active
for idx, x in enumerate(list(restaurant_data[0])+['1', '2', '3', '4', '5'], 1):
    sheet.cell(row=1, column=idx, value=x)

wb.save(excel_full_path)



import requests
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager


import time
import re

options = webdriver.ChromeOptions()
# options.add_argument("headless")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")
options.add_argument("--blink-settings=imagesEnabled=false")
w, h = 1920, 1080
driver = webdriver.Chrome(options=options)
driver.set_window_size(w, h)
driver.set_script_timeout(5)

wait = WebDriverWait(driver, 10)

total_count = 0
for idx, restaurant_row in enumerate(restaurant_data[1:], 2):
    name = restaurant_row[21]
    print(f'try {idx}! {name}:[{sheet.cell(row=idx, column=len(restaurant_data[0])+1).value}]')
    
    if sheet.cell(row=idx, column=len(restaurant_data[0])+1).value not in ['None', None, '',]:
        continue
    for idx2, x in enumerate(list(restaurant_data[idx-1]), 1):
        sheet.cell(row=idx, column=idx2, value=x)

    url = f'https://map.kakao.com/'
    driver.get(url)

    search_box = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="search.keyword.query"]')))
    search_box.send_keys(name)
    # search_button = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="search.keyword.submit"]')))
    # search_button.click()
    search_box.send_keys(Keys.ENTER)

    xpath = '//*[@id="info.search.place.list"]/li/div[3]/span'
    try:
        sidebar = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
    except:
        print(f'can\'t find sidebar... - {name} [{idx}]')

    xpath_format = '//*[@id="info.search.place.list"]/li[{}]/div[3]/span'
    xpath_idx = 1

    category_dict = dict()
    while True:
        try:
            now_xpath = xpath_format.format(xpath_idx)
            category = wait.until(EC.presence_of_element_located((By.XPATH, now_xpath))).text
            if category not in category_dict:
                category_dict[category] = 0
            category_dict[category] += 1
            xpath_idx += 1
        except:
            break

    sorted_dict = list(sorted(category_dict.items(), key=lambda x: x[1], reverse=True))
    # if len(sorted_dict) < 5:
    #     for _ in range(len(sorted_dict), 5):
    #         sorted_dict.append(["not found", 0])
    for idx3, (k, v) in enumerate(sorted_dict, 1):
        sheet.cell(row=idx, column=len(restaurant_data[0])+idx3, value=f"{k} ({v})")
        print(idx, len(restaurant_data[0])+idx3)
    if len(sorted_dict) == 0:
        sheet.cell(row=idx, column=len(restaurant_data[0])+1, value=f"-")
        print("-")
    wb.save(excel_full_path)
    print(f'at name:{name}, category is {sorted_dict}[{idx}]')
    time.sleep(1)