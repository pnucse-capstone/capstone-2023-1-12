import openpyxl
from openpyxl import Workbook
from os.path import dirname, realpath, exists

# Open Brand
org_path = dirname(realpath(__file__).replace("\\", "/"))

def get_category_data(excel_file_path):
    excel_full_path = '/'.join([org_path, excel_file_path])
    wb = openpyxl.load_workbook(excel_full_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data



restaurant_data = get_category_data('data.xlsx')

excel_full_path = '/'.join([org_path, 'search.xlsx'])
if exists(excel_full_path):
    wb = openpyxl.load_workbook(excel_full_path)
else:
    wb = Workbook()
sheet = wb.active
for idx, x in enumerate(list(restaurant_data[0])+['naver(name)', 'naver(address)', 'kakao', 'google'], 1):
    sheet.cell(row=1, column=idx, value=x)

wb.save(excel_full_path)



import requests
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager


import time
import re

options = webdriver.ChromeOptions()
# options.add_argument("headless")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")
options.add_argument("--blink-settings=imagesEnabled=false")
w, h = 1920, 1080
driver = webdriver.Chrome(options=options)
driver.set_window_size(w, h)
driver.set_script_timeout(5)

wait = WebDriverWait(driver, 10)

total_count = 0
for idx, restaurant_row in enumerate(restaurant_data[1:], 2):
    print(f'try {idx}!')
    if idx > 10: break
    name = restaurant_row[21]
    
    if sheet.cell(row=idx, column=len(restaurant_data)+1).value not in ['None', None, '']:
        continue
    for idx2, x in enumerate(list(restaurant_data[idx]), 1):
        sheet.cell(row=idx, column=idx2, value=x)

    url = f'https://map.naver.com/p/search/{name}?c=13.29,0,0,0,dh'
    driver.get(url)
    time.sleep(5)
    
    xpath = '//*[@id="_pcmap_list_scroll_container"]'
    try:
        sidebar = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
    except:
        print(f'can\'t find sidebar...[{idx}]')
        try:
            category = driver.find_element(By.XPATH, '//*[@id="_title"]/span[2]').text
            sheet.cell(row=idx, column=len(restaurant_data)+1, value=category)
            wb.save(excel_full_path)
            print(f'at name:{name}, category is {category}[{idx}]')
            continue
        except:
            sheet.cell(row=idx, column=len(restaurant_data)+1, value='not found')
            wb.save(excel_full_path)
            print('not matching...')
            continue

    xpath_format = '//*[@id="_pcmap_list_scroll_container"]/ul/li[{}]'

    xpath_idx = 1

    while True:
        sidebar_height = sidebar.size['height']
        driver.execute_script(f"arguments[0].scrollTo(0, {sidebar_height});", sidebar)
        
        # 잠시 대기하여 스크롤이 완료되기를 기다림 (필요에 따라 대기 시간 조절)
        time.sleep(2)
        
        # 스크롤을 더 이상 내릴 수 없으면 반복 종료
        if sidebar_height == sidebar.size['height']:
            print('scroll')
            break
    while True:
        try:
            now_xpath = xpath_format.format(xpath_idx)
            li_element = driver.find_element(By.XPATH, now_xpath)
            address = driver.find_element(By.XPATH, now_xpath+'/div[1]/div/div/div/span/a/span[1]').text
            if address in '금정구':
                category = driver.find_element(By.XPATH, now_xpath+'/div[1]/div/a[1]/div/div/span[2]').text
                sheet.cell(row=idx, column=len(restaurant_data)+1, value=category)
                wb.save(excel_full_path)
                print(f'at name:{name}, category is {category}[{idx}]')
                break
            
            xpath_idx += 1
        except:
            try:
                next_button_xpath = '//*[@id="app-root"]/div/div[2]/div[2]/a[5]/svg'
                next_button = driver.find_element(By.XPATH, next_button_xpath)
                print('next page')
            except:
                sheet.cell(row=idx, column=len(restaurant_data)+1, value='not found')
                wb.save(excel_full_path)
                print('not matching...')
                break
    time.sleep(2)