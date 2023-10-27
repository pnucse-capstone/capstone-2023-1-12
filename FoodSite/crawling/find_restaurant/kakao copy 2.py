import openpyxl
from openpyxl import Workbook
from os.path import dirname, realpath, exists

# Open Brand
org_path = dirname(realpath(__file__).replace("\\", "/"))

def get_category_data(excel_file_path):
    excel_full_path = '/'.join([org_path, excel_file_path])
    wb = openpyxl.load_workbook(excel_full_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data



restaurant_data = get_category_data('search_kakao 2.xlsx')

excel_full_path = '/'.join([org_path, 'search_kakao 3.xlsx'])
if exists(excel_full_path):
    wb = openpyxl.load_workbook(excel_full_path)
else:
    wb = Workbook()
sheet = wb.active

wb.save(excel_full_path)



import re
# from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
# from webdriver_manager.chrome import ChromeDriverManager


import time
import re

options = webdriver.ChromeOptions()
# options.add_argument("headless")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")
options.add_argument("--blink-settings=imagesEnabled=false")
w, h = 1920, 1080
driver = webdriver.Chrome(options=options)
driver.set_window_size(w, h)
driver.set_script_timeout(5)

wait = WebDriverWait(driver, 5)


recent_row = sheet.cell(row=sheet.max_row, column=1).value
excel_idx = sheet.max_row

# import re

for idx, restaurant_row in enumerate(restaurant_data[1:], 2):
    if idx < recent_row:
        continue

    name = restaurant_row[21]
    print(f'try {idx}! {name}')
    
    url = restaurant_row[50]
    print(url)
    if url in ["", None]: continue
    
    driver.get(url)

    try:
        menu_box = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="mArticle"]/div[3]/ul')))
    except:
        print("can't find!!!")
        
    
    menulist_selector = '#mArticle > div.cont_menu > ul'

    menu_name_selector  = '#mArticle > div.cont_menu > ul > li:nth-child({}) > div > span'
    # menu_name_selector  = #mArticle > div.cont_menu > ul > li:nth-child(1) > div > span
    menu_price_selector = '#mArticle > div.cont_menu > ul > li:nth-child({}) > div > em.price_menu'
    #mArticle > div.cont_menu > a

    cnt = 1
    while True:
        try:
            print("---" + driver.find_element(By.CSS_SELECTOR, '#mArticle > div.cont_menu > a > span.open_txt').text)
            if driver.find_element(By.CSS_SELECTOR, '#mArticle > div.cont_menu > a > span.open_txt').text == "메뉴 더보기":
                driver.find_element(By.CSS_SELECTOR, '#mArticle > div.cont_menu > a').click()
            else:
                break
        except:
            break

            
    time.sleep(1)
    while True:
        # menu 찾기
        try:
            menu = driver.find_element(By.CSS_SELECTOR, menu_name_selector.format(cnt)).text
            try:
                price = driver.find_element(By.CSS_SELECTOR, menu_price_selector.format(cnt)).text
                pass
            except:
                # price 없으면
                price = '0'
        except:
            # menu 없으면
            break

        sheet.cell(row=excel_idx, column=1, value=idx)
        sheet.cell(row=excel_idx, column=2, value=menu)
        sheet.cell(row=excel_idx, column=3, value=price)

        print(f"[{idx}] {menu} = {price}")
        wb.save(excel_full_path)

        cnt += 1
        excel_idx += 1

    time.sleep(1)