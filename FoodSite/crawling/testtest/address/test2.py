import openpyxl
from openpyxl import Workbook
from os.path import dirname, realpath, exists

# youtube excel
org_path = dirname(realpath(__file__).replace("\\", "/"))

youtube_excel_path = '/'.join([org_path, '부산 금정구 식당 - total - 주소 분리 시도.xlsx'])

youtube_wb      = openpyxl.load_workbook(youtube_excel_path)
youtube_sheet   = youtube_wb.active

# restaurant excel
restaurant_excel_path = '/'.join([org_path, '식당 목록.xlsx'])

restaurant_wb         = openpyxl.load_workbook(restaurant_excel_path)
restaurant_sheet      = restaurant_wb["Sheet"]

# result excel
result_path = '/'.join([org_path, '식당명 대비 후보 식당.xlsx'])
if exists(result_path):
    result_wb = openpyxl.load_workbook(result_path)
else:
    result_wb = Workbook()
result_sheet = result_wb.active

# create restaurant dictionary
restaurants = []
sample = {
    "idx"       : 0,    # excel에서의 row number
    "name1"     : "",   # 가게 이름
    "name2"     : "",   # kakao에서 찾은 가게 이름
    "address1"  : "",   # 소재지주소
    "address2"  : "",   # 도로명주소
    "address3"  : "",   # kakao 주소
}

from address import Korea
from korea_.busan import Busan
from korea_.busan_.guemjeong import Guemjeong

korea = Korea()

RESTAURANT_NAME = 22
FULL_ADDRESS = 19
ROAD_ADDRESS = 20
X_COORDINATE = 27
Y_COORDINATE = 28

KAKAO_RESTAURANT_NAME = 48
KAKAO_ADDRESS = 50
CATEGORY_DEFAULT = 49

for row in range(2, restaurant_sheet.max_row+1):
    if restaurant_sheet.cell(row=row, column=CATEGORY_DEFAULT).value in  ["", None]:
        continue

    data = {
        "idx"       : row,          # excel에서의 row number
        "name1"     : restaurant_sheet.cell(row=row, column=RESTAURANT_NAME).value, # 가게 이름
        "name2"     : restaurant_sheet.cell(row=row, column=KAKAO_RESTAURANT_NAME).value,   # kakao에서 찾은 가게 이름
        "address1"  : restaurant_sheet.cell(row=row, column=FULL_ADDRESS).value,    # 소재지주소
        "address2"  : restaurant_sheet.cell(row=row, column=ROAD_ADDRESS).value,    # 도로명주소
        "address3"  : restaurant_sheet.cell(row=row, column=KAKAO_ADDRESS).value,   # kakao 주소
    }
    data2 = {}
    for key, value in data.items():
        data2[key] = value if value else ""

    restaurants.append(data2)
    # break

import re
import Levenshtein

for row in range(2, youtube_sheet.max_row+1):
    if row > 10:
        break
    addr = youtube_sheet.cell(row=row, column=11).value
    full = youtube_sheet.cell(row=row, column=8).value
    if not addr:
        continue
    data1, data2 = korea.calculate(addr)
    modified_texts = [x.strip() for x in re.sub(r'(\d+)', r' \1', " ".join(data2[0])).split()]
    print("org::", full)
    print(modified_texts)

    res = []
    for restaurant in restaurants:
        address = restaurant['address1'] + " " \
                + restaurant['address2'] + " " \
                + restaurant['address3']
        isbreak = False
        for x in modified_texts:
            if x not in address:
                isbreak = True
                break
        if not isbreak: res.append(restaurant["name1"])
        
    print(korea.calculate(addr))
    print("\nres: ", res, "\n\n\n\n")

    print("Levenshtein 결과: ")
    res2 = []
    for r in res:
        distance = Levenshtein.distance(full, r)
        res2.append((distance, r))
    print(sorted(res2))
        


    

result_wb.save(result_path)
