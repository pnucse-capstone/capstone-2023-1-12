import openpyxl
from openpyxl import Workbook
from os.path import dirname, realpath, exists

# Open Brand
org_path = dirname(realpath(__file__).replace("\\", "/"))

excel_full_path = '/'.join([org_path, '부산 금정구 식당 - total.xlsx'])
excel_full_path2 = '/'.join([org_path, '부산 금정구 식당 - total - 주소 분리 시도.xlsx'])
if exists(excel_full_path):
    wb = openpyxl.load_workbook(excel_full_path)
else:
    wb = Workbook()
sheet = wb.active

for idx, x in enumerate(['idx', 'name', 'view_count', 'channel', 'address', 'real_address', 'hash_tag', 'description'], 1):
    sheet.cell(row=1, column=idx, value=x)

wb.save(excel_full_path2)

from address import Korea
from korea_.busan import Busan
from korea_.busan_.guemjeong import Guemjeong

korea = Korea()




for row in range(1, sheet.max_row):
    # if sheet.cell(row=row+1, column=10).value not in ["None", None, '']:
    #     continue
    if sheet.cell(row=row+1, column=8).value in ["None", None, '']:
        continue

    description = sheet.cell(row=row+1, column=8).value
    is_0 = 0
    for x in ["1.", "2.", "3."]:
        if x not in description:
            is_0 = 1
            break

    print(not is_0)

    sheet.cell(row=row+1, column=12, value=1-is_0)

    address = sheet.cell(row=row+1, column=8).value

    my_address = korea.calculate(address)
    my_address2 = korea.calculate(address, want_class=[Korea, Busan, Guemjeong])
    # my_address = korea.check(address)
    print(my_address)
    sheet.cell(row=row+1, column=10, value=my_address)
    sheet.cell(row=row+1, column=11, value=my_address2)

    wb.save(excel_full_path2)

    # break
    if row > 30: break