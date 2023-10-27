import openpyxl
from openpyxl import Workbook
from os.path import dirname, realpath, exists

# Open Brand
org_path = dirname(realpath(__file__).replace("\\", "/"))

excel_full_path = '/'.join([org_path, '부산 금정구 식당.xlsx'])
excel_full_path2 = '/'.join([org_path, '부산 금정구 식당 - 수정본.xlsx'])
if exists(excel_full_path):
    wb = openpyxl.load_workbook(excel_full_path)
else:
    wb = Workbook()
sheet = wb.active

for idx, x in enumerate(['idx', 'name', 'view_count', 'channel', 'address', 'real_address', 'hash_tag', 'description'], 1):
    sheet.cell(row=1, column=idx, value=x)

wb.save(excel_full_path2)






import requests
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager


import time
import re

options = webdriver.ChromeOptions()
# options.add_argument("headless")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")
options.add_argument("--blink-settings=imagesEnabled=false")
w, h = 1920, 1080
driver = webdriver.Chrome(options=options)
driver.set_window_size(w, h)
driver.set_script_timeout(5)

wait = WebDriverWait(driver, 10)

count = 0


def address_split(x):
    if 'shorts' in x:
        return x.split('shorts/')[1]
    x = x.split('watch?v=')[1]
    if '&pp=' in x:
        return x.split('&pp=')[0]

def find_elememt_until(wait, xpath) -> WebDriverWait:
    return wait.until(EC.presence_of_element_located((By.XPATH, xpath)))


def scroll_down():
    # page_height = driver.execute_script("return document.body.scrollHeight")
    # driver.execute_script(f"window.scrollTo(0, {page_height});")

    footer = driver.find_element(By.TAG_NAME, "body")
    footer.send_keys(Keys.END)
    footer.send_keys(Keys.END)
    footer.send_keys(Keys.END)
    time.sleep(5)



my_xpath = '//*[@id="contents"]/ytd-video-renderer[{}]'
url = f'https://www.youtube.com/'
driver.get(url)

internal_count = 1

total_count = 0
for row in range(1, sheet.max_row):
    if sheet.cell(row=row+1, column=6).value not in ["None", None, '']:
        continue

    address = address_split(sheet.cell(row=row+1, column=5).value)
    url = f'https://www.youtube.com/watch?v={address}'
    driver.get(url)
    
    button = '//*[@id="expand"]'
    try:
        wait.until(EC.presence_of_element_located((By.XPATH, button))).click()
    except:
        pass
    
    x1 = '//*[@id="info-container"]'
    x2 = '//*[@id="description-inline-expander"]/yt-attributed-string'

    hash_tag = ['#'+x.strip() for x in wait.until(EC.presence_of_element_located((By.XPATH, x1))).text.split('#')[1:]]
    description = wait.until(EC.presence_of_element_located((By.XPATH, x2))).text

    for hash2 in description.split():
        hash2 = hash2.split('\n')[0].strip()
        if hash2[0] == '#' and hash2 not in hash_tag:
            hash_tag.append(hash2)

    print(hash_tag)
    print(description)

    sheet.cell(row=row+1, column=6, value=' '.join(address))
    sheet.cell(row=row+1, column=7, value=' '.join(hash_tag))
    sheet.cell(row=row+1, column=8, value=description)

    wb.save(excel_full_path2)