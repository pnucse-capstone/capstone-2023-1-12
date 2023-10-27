import openpyxl
import pyperclip
from openpyxl import Workbook
from os.path import dirname, realpath, exists

# Open Brand
org_path = dirname(realpath(__file__).replace("\\", "/"))

excel_full_path = '/'.join([org_path, '부산 금정구 식당 - total.xlsx'])
excel_full_path2 = '/'.join([org_path, '부산 금정구 식당 - total - gpt.xlsx'])
if exists(excel_full_path):
    wb = openpyxl.load_workbook(excel_full_path)
else:
    wb = Workbook()
sheet = wb.active

wb.save(excel_full_path2)

default_text = '''

'''
question_text = '''아래에 내가 반복된 내용을 올릴거야. 이들은 Youtube 영상의 description이야. 이 영상이 식당의 정보를 담고 있는지를 알고 싶어. 따라서 아래의 동작을 해줬으면 해.
이 description은 하나의 식당에 대한 정보가 있을수도 있고, 여러개의 식당에 대한 정보가 있을수도 있어. 또는 식당에 대한 내용이 아닐 수도 있지.

- 내가 지정한 양식 이외의 내용은 절대 작성하지 마. 오로지 아래의 양식에 맞는 내용만 작성해.
- 이 description이 식당들에 대한 정보인지 알고싶어. 만약 식당이 아니라면, False만 출력해. 오직 그것만 있으면 돼.
- google map link 등의 link 내용은 주소가 아니야. 이들에 제외한 것에서 주소 찾기를 시도해.
- 식당이 하나의 식당의 정보만 가지고 있더라도, numbering을 할 거야.
- numbering 양식은 아래와 같아. index는 여기서 numbering 할 숫자야. []로 감싼 부분을 적절한 값으로 채워주면 돼.
    [index]: [식당 이름]
    주소: [식당의 주소]
'''



import pyautogui
import time

chat_box = (-1197, 979)
default_mouse = (-1102, 779)


for idx in range(1, sheet.max_row):
    description = sheet.cell(row=idx+1, column=8).value

    # my_qyestion_text = question_text.format(description)
    my_qyestion_text = description
    pyperclip.copy(my_qyestion_text)
    pyautogui.moveTo(*chat_box, duration=1)
    time.sleep(1)
    pyautogui.click()
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.hotkey('enter')

    time.sleep(20)

    # pyautogui.moveTo(*default_mouse, duration=1)
    # time.sleep(1)
    # pyautogui.click()
    # time.sleep(1)
    # pyautogui.hotkey('ctrl', 'a')
    # time.sleep(1)
    # pyautogui.hotkey('ctrl', 'c')

    # clipboard_text = pyperclip.paste()
    # want_text = clipboard_text.split(my_qyestion_text)[-1]

    # print(want_text)

    # sheet.cell(row=idx+1, column=9, value=want_text)
    # wb.save(excel_full_path2)