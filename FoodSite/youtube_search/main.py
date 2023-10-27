import openpyxl
from openpyxl import Workbook
from os.path import dirname, realpath, exists

# Open Brand
org_path = dirname(realpath(__file__).replace("\\", "/"))

excel_full_path = '/'.join([org_path, 'youtube_search.xlsx'])
if exists(excel_full_path):
    wb = openpyxl.load_workbook(excel_full_path)
else:
    wb = Workbook()
sheet = wb.active

for idx, x in enumerate(['idx', 'name', 'view_count', 'channel', 'address'], 1):
    sheet.cell(row=1, column=idx, value=x)

wb.save(excel_full_path)






import requests
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service


import time
import re

options = webdriver.ChromeOptions()
# options.add_argument("headless")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")
options.add_argument("--blink-settings=imagesEnabled=false")
w, h = 1920, 1080

service = Service(excutable_path='/chromedriver_linux')

driver = webdriver.Chrome(service=service, options=options)
driver.set_window_size(w, h)
driver.set_script_timeout(5)

wait = WebDriverWait(driver, 10)

count = 0




def find_elememt_until(wait, xpath) -> WebDriverWait:
    return wait.until(EC.presence_of_element_located((By.XPATH, xpath)))


def scroll_down():
    # page_height = driver.execute_script("return document.body.scrollHeight")
    # driver.execute_script(f"window.scrollTo(0, {page_height});")

    footer = driver.find_element(By.TAG_NAME, "body")
    footer.send_keys(Keys.END)
    footer.send_keys(Keys.END)
    footer.send_keys(Keys.END)
    time.sleep(5)



my_xpath = '//*[@id="contents"]/ytd-video-renderer[{}]'
url = f'https://www.youtube.com/results?search_query=부산+금정구+식당'
driver.get(url)

internal_count = 1

while True:

    video_elements = driver.find_elements(By.XPATH, '//*[@id="contents"]/ytd-video-renderer')
    for x in range(count, len(video_elements)):
        video_element = video_elements[x]
    
        video = {}
        video['idx'] = internal_count
        video["name"] = video_element.find_element(By.CSS_SELECTOR, '#video-title > yt-formatted-string').text
        video["view_count"] = video_element.find_element(By.CSS_SELECTOR, '#metadata-line > span:nth-child(3)').text
        video["channel"] = video_element.find_element(By.CSS_SELECTOR, '#text').text
        video["address"] = video_element.find_element(By.CSS_SELECTOR, '#video-title').get_attribute("href")

        for idx, val in enumerate(video.values(), 1):
            sheet.cell(row=internal_count+1, column=idx, value=val)
        wb.save(excel_full_path)

        internal_count += 1

        print(video)

    scroll_down()
    if count == len(video_elements):
        break
    count = len(video_elements)

    

# for idx, restaurant_row in enumerate(restaurant_data[1:], 2):
#     name = restaurant_row[21]
#     print(f'try {idx}! {name}:[{sheet.cell(row=idx, column=len(restaurant_data[0])+1).value}]')
    
#     if sheet.cell(row=idx, column=len(restaurant_data[0])+1).value not in ['None', None, '',]:
#         continue
#     for idx2, x in enumerate(list(restaurant_data[idx-1]), 1):
#         sheet.cell(row=idx, column=idx2, value=x)

#     url = f'https://map.kakao.com/'
#     driver.get(url)

#     search_box = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="search.keyword.query"]')))
#     search_box.send_keys(name)
#     # search_button = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="search.keyword.submit"]')))
#     # search_button.click()
#     search_box.send_keys(Keys.ENTER)

#     xpath = '//*[@id="info.search.place.list"]/li/div[3]/span'
#     try:
#         sidebar = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
#     except:
#         print(f'can\'t find sidebar... - {name} [{idx}]')

#     xpath_format = '//*[@id="info.search.place.list"]/li[{}]/div[3]/span'
#     xpath_idx = 1

#     category_dict = dict()
#     while True:
#         try:
#             now_xpath = xpath_format.format(xpath_idx)
#             category = wait.until(EC.presence_of_element_located((By.XPATH, now_xpath))).text
#             if category not in category_dict:
#                 category_dict[category] = 0
#             category_dict[category] += 1
#             xpath_idx += 1
#         except:
#             break

#     sorted_dict = list(sorted(category_dict.items(), key=lambda x: x[1], reverse=True))
#     # if len(sorted_dict) < 5:
#     #     for _ in range(len(sorted_dict), 5):
#     #         sorted_dict.append(["not found", 0])
#     for idx3, (k, v) in enumerate(sorted_dict, 1):
#         sheet.cell(row=idx, column=len(restaurant_data[0])+idx3, value=f"{k} ({v})")
#         print(idx, len(restaurant_data[0])+idx3)
#     if len(sorted_dict) == 0:
#         sheet.cell(row=idx, column=len(restaurant_data[0])+1, value=f"-")
#         print("-")
#     wb.save(excel_full_path)
#     print(f'at name:{name}, category is {sorted_dict}[{idx}]')
#     time.sleep(1)