from django.contrib  import admin
from app_model.database.models import *
from app_model.extra.models    import *
import openpyxl
from math import isnan
import os   
from django.conf import settings
import re
# Register your models here.

def get_or_create_service_idx(model, service_name, **extra):
    try:
        # 이미 해당 이름으로 레코드가 있는지 확인
        service = model.objects.get(name=service_name)
        if service.idx == None: raise ValueError
        return service
    except model.DoesNotExist:
        # 레코드가 없으면 새로 생성
        new_service = model(name=service_name, **extra)
        new_service.save()
        return new_service


def import_data(modeladmin, request, queryset):
    # 지정된 상대 경로의 data.xlsx 파일을 읽어서 DataFrame으로 가져옴
    file_name = 'data/data.xlsx'
    file_path = os.path.join(settings.BASE_DIR, file_name)

    wb = openpyxl.load_workbook(file_path)

    nan_to_zero_2 = lambda x: 0 if isnan(x) else x
    nan_to_zero = lambda x: 0 if x == None else nan_to_zero_2(float(x.strip()))
    replace_none_to_blank = lambda x: "" if x in ["", None] else x

    num = {
        "total" :0,
        "update":0,
        "new"   :0,
    }

    RESTAURANT_NAME = 22
    FULL_ADDRESS = 19
    ROAD_ADDRESS = 20
    X_COORDINATE = 27
    Y_COORDINATE = 28

    for sheet_name in wb.sheetnames:
        if sheet_name == 'Main': continue
        ws = wb[sheet_name]
        for row in range(2, ws.max_row+1):
            # 소재지주소 19
            # 도로명주소 20
            # 가게명 22
            # x 27
            # y 28
            # 카테고리    sheet_name
            num["total"] += 1

            try:
                obj = Restaurant.objects.filter(
                    name=ws.cell(row=row, column=RESTAURANT_NAME).value,
                    x_coordinate=round(nan_to_zero(ws.cell(row=row, column=X_COORDINATE).value), 9),
                    y_coordinate=round(nan_to_zero(ws.cell(row=row, column=Y_COORDINATE).value), 9),
                ).first()

                if (obj.full_address, obj.road_address) ==\
                   (replace_none_to_blank(ws.cell(row=row, column=FULL_ADDRESS).value), replace_none_to_blank(ws.cell(row=row, column=ROAD_ADDRESS).value)):
                    continue
            except Exception as e:
                # 데이터가 없으면 새로운 객체 생성 후 저장
                Restaurant.objects.create(
                    name=ws.cell(row=row, column=RESTAURANT_NAME).value,
                    x_coordinate=round(nan_to_zero(ws.cell(row=row, column=X_COORDINATE).value), 9),
                    y_coordinate=round(nan_to_zero(ws.cell(row=row, column=Y_COORDINATE).value), 9),
                    full_address=replace_none_to_blank(ws.cell(row=row, column=FULL_ADDRESS).value),
                    road_address=replace_none_to_blank(ws.cell(row=row, column=ROAD_ADDRESS).value),
                    service=get_or_create_service_idx(RestaurantServiceName, sheet_name, ),
                )
                num["new"] += 1
        
    Order.objects.create(
        name="Update data",
        description="\n".join([f"{k}: {v}" for k, v in num.items()])
    )

import_data.short_description = "Import data from data.xlsx"


def import_data_2(modeladmin, request, queryset):
    # 지정된 상대 경로의 data.xlsx 파일을 읽어서 DataFrame으로 가져옴
    restaurant_file_name = 'data/data_restaurant.xlsx'
    restaurant_file_path = os.path.join(settings.BASE_DIR, restaurant_file_name)
    menu_file_name = 'data/data_menu.xlsx'
    menu_file_path = os.path.join(settings.BASE_DIR, menu_file_name)

    wb_restaurant = openpyxl.load_workbook(restaurant_file_path)
    wb_menu = openpyxl.load_workbook(menu_file_path)

    nan_to_zero_2 = lambda x: 0 if isnan(x) else x
    nan_to_zero = lambda x: 0 if x == None else nan_to_zero_2(float(x.strip()))
    replace_none_to_blank = lambda x: "" if x in ["", None] else x

    num = {
        "total" :0,
        "update":0,
        "new"   :0,
    }

    restaurant_category = {}
    sheet_category = wb_restaurant["시트2"]
    english_name = {}
    
    idx = 2
    while True:
        v = sheet_category.cell(row=idx, column=8).value
        if v in ["None", None, ""]:
            break
        english_name[v] = sheet_category.cell(row=idx, column=9).value
        idx += 1

    for i in range(1, sheet_category.max_row):
        v = sheet_category.cell(row=i+1, column=2).value
        restaurant_category[sheet_category.cell(row=i+1, column=1).value] = v
        get_or_create_service_idx(RestaurantServiceName, v, service=english_name[v])

    RESTAURANT_NAME = 22
    FULL_ADDRESS = 19
    ROAD_ADDRESS = 20
    X_COORDINATE = 27
    Y_COORDINATE = 28
    CATEGORY_DEFAULT = 49

    sheet_restaurant = wb_restaurant["Sheet"]
    sheet_menu = wb_menu["Sheet"]

    menu_idx = 2

    for idx in range(1+1, sheet_restaurant.max_row+1):
        if sheet_restaurant.cell(row=idx, column=CATEGORY_DEFAULT).value not in restaurant_category:
            continue

        num["total"] += 1

        try:
            obj = Restaurant.objects.filter(
                name=sheet_restaurant.cell(row=idx, column=RESTAURANT_NAME).value,
                x_coordinate=round(nan_to_zero(sheet_restaurant.cell(row=idx, column=X_COORDINATE).value), 9),
                y_coordinate=round(nan_to_zero(sheet_restaurant.cell(row=idx, column=Y_COORDINATE).value), 9),
            ).first()

            if (obj.full_address, obj.road_address) ==\
                (replace_none_to_blank(sheet_restaurant.cell(row=idx, column=FULL_ADDRESS).value), replace_none_to_blank(sheet_restaurant.cell(row=idx, column=ROAD_ADDRESS).value)):
                pass
        except Exception as e:
            # 데이터가 없으면 새로운 객체 생성 후 저장
            Restaurant.objects.create(
                name=sheet_restaurant.cell(row=idx, column=RESTAURANT_NAME).value,
                x_coordinate=round(nan_to_zero(sheet_restaurant.cell(row=idx, column=X_COORDINATE).value), 9),
                y_coordinate=round(nan_to_zero(sheet_restaurant.cell(row=idx, column=Y_COORDINATE).value), 9),
                full_address=replace_none_to_blank(sheet_restaurant.cell(row=idx, column=FULL_ADDRESS).value),
                road_address=replace_none_to_blank(sheet_restaurant.cell(row=idx, column=ROAD_ADDRESS).value),
                service=get_or_create_service_idx(RestaurantServiceName, restaurant_category[sheet_restaurant.cell(row=idx, column=CATEGORY_DEFAULT).value]),
            )
            num["new"] += 1
            obj = Restaurant.objects.filter(
                name=sheet_restaurant.cell(row=idx, column=RESTAURANT_NAME).value,
                x_coordinate=round(nan_to_zero(sheet_restaurant.cell(row=idx, column=X_COORDINATE).value), 9),
                y_coordinate=round(nan_to_zero(sheet_restaurant.cell(row=idx, column=Y_COORDINATE).value), 9),
            ).first()

        while menu_idx <= sheet_menu.max_row:
            while menu_idx <= sheet_menu.max_row and sheet_menu.cell(row=menu_idx, column=1).value < idx:
                menu_idx += 1
            if sheet_menu.cell(row=menu_idx, column=1).value == idx:
                myprice = re.sub(r',', '', str(sheet_menu.cell(row=menu_idx, column=3).value))
                myprice = int(myprice) if myprice.isdigit() else 0
                try:
                    obj2 = RestaurantMenu.objects.filter(
                        restaurant=obj,
                        name=sheet_menu.cell(row=menu_idx, column=2).value,
                        price=myprice,
                    ).first()
                    if obj2 is None:
                        raise Exception
                except:
                    RestaurantMenu.objects.create(
                        restaurant=obj,
                        name=sheet_menu.cell(row=menu_idx, column=2).value,
                        price=myprice,
                    )
                    print(f"create menu: {obj} - {sheet_menu.cell(row=menu_idx, column=2).value}")
                menu_idx += 1
            else:
                break
        
    Order.objects.create(
        name="Update data",
        description="\n".join([f"{k}: {v}" for k, v in num.items()])
    )

import_data_2.short_description = "Import data from kakao.xlsx"



def delete_all_data(self, request, queryset):
    # 선택한 모델의 모든 데이터를 삭제합니다.
    Restaurant.objects.all().delete()
    self.message_user(request, "선택한 모델의 모든 데이터가 삭제되었습니다.")

delete_all_data.short_description = "선택한 모델의 모든 데이터 삭제"


class YourModelAdmin(admin.ModelAdmin):
    actions = [import_data, import_data_2, delete_all_data]
    
class RestaurantAdmin(admin.ModelAdmin):
    search_fields = ['name']

class RestaurantVideoAdmin(admin.ModelAdmin):
    raw_id_fields = ('restaurant',)
    # search_fields = ('restaurant__name',)
    
    def restaurant_name(self, obj):
        return obj.restaurant.name

admin.site.register(Restaurant, RestaurantAdmin)

admin.site.register(RestaurantMenu)
admin.site.register(RestaurantVideo, RestaurantVideoAdmin)
admin.site.register(RestaurantAdditionalInfo)
admin.site.register(RestaurantServiceName)



admin.site.register(Order, YourModelAdmin)